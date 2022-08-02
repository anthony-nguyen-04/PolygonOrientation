import openpyxl
import numpy as np

def checkDirection(zCross):
    if zCross > 0:
        return "CLOCKWISE"
    elif zCross < 0:
        return "COUNTER-CLOCKWISE"
    else:
        return "N/A"

location = "PolygonOrientation.xlsx"

wb = openpyxl.load_workbook(location)
sheet = wb["shape2"]

points = {}

for count, row in enumerate(sheet.iter_rows(max_row = sheet.max_row - 1)):
    coordinate = (row[1].value, row[2].value)
    #print("\draw[fill=black] (%s, %s) circle (2pt);" % (row[1].value, row[2].value))
    print("\\node at (%s, %s) {P%s};" % (row[1].value, row[2].value+0.5, count+1))
    points[count] = coordinate



bottomPoint = sorted(points.items(), key = lambda x: x[1][1])[0]
bottomIndex = bottomPoint[0]

(xCenter, yCenter) = points[bottomIndex]
(xBefore, yBefore) = points[(bottomIndex - 1) % len(points)]
(xAfter, yAfter) = points[(bottomIndex + 1) % len(points)]

vectorBefore = (xBefore - xCenter, yBefore - yCenter)
vectorAfter = (xAfter - xCenter, yAfter - yCenter)

vectorBefore = np.asarray(vectorBefore)
vectorAfter = np.asarray(vectorAfter)

cross = np.cross(vectorBefore, vectorAfter)

print(cross)
print(checkDirection(cross))
